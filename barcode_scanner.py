#!/usr/bin/env python3
"""
Enhanced Barcode Scanner Project with Box Upload:
- Improved error handling and resilience
- Better configuration validation
- Barcode input timeout and validation
- Robust upload system with worker thread and queue
- Structured logging
- Only uploads files with actual scan data
"""

import os
import time
import schedule
import threading
import logging
import queue
from datetime import datetime, timedelta
from typing import Optional, Dict, Any
import json
from pathlib import Path

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from dotenv import load_dotenv
import keyboard

# Import Box SDK JWT classes
from boxsdk import JWTAuth, Client
from boxsdk.exception import BoxAPIException

# --- Configuration and Logging Setup ---
class Config:
    def __init__(self):
        # Load environment variables
        load_dotenv()
        
        # Set base directory relative to project root
        self.base_dir = Path(os.getenv("BASE_DIR", "./data")).resolve()
        
        # Required environment variables
        self.box_config_path = self._get_required_env("BOX_CONFIG_PATH")
        self.target_folder_id = self._get_required_env("TARGET_FOLDER_ID")
        self.collaborator_email = self._get_required_env("COLLABORATOR_EMAIL")
        
        # Optional settings with defaults
        self.barcode_timeout = int(os.getenv("BARCODE_TIMEOUT", "5"))
        self.max_retries = int(os.getenv("MAX_RETRIES", "3"))
        self.retry_delay = int(os.getenv("RETRY_DELAY", "5"))
        self.log_level = os.getenv("LOG_LEVEL", "INFO")
        
        self._validate_config()
    
    def _get_required_env(self, key: str) -> str:
        value = os.getenv(key)
        if not value:
            raise ValueError(f"Required environment variable {key} is not set")
        return value
    
    def _validate_config(self):
        """Validate configuration files and paths exist"""
        # Create base directory if it doesn't exist
        self.base_dir.mkdir(parents=True, exist_ok=True)
        
        # Create logs directory
        logs_dir = Path("./logs")
        logs_dir.mkdir(exist_ok=True)
        
        # Validate Box config file
        box_config_path = Path(self.box_config_path)
        if not box_config_path.exists():
            raise FileNotFoundError(f"Box config file not found: {box_config_path}")
        
        # Validate Box config JSON structure
        try:
            with open(box_config_path, 'r') as f:
                config_data = json.load(f)
                required_keys = ['boxAppSettings', 'enterpriseID']
                if not all(key in config_data for key in required_keys):
                    raise ValueError("Invalid Box config file structure")
        except json.JSONDecodeError:
            raise ValueError("Box config file contains invalid JSON")

# Set up structured logging
def setup_logging(log_level: str = "INFO"):
    log_dir = Path("./logs")
    log_dir.mkdir(exist_ok=True)
    log_file = log_dir / "barcode_scanner.log"
    
    # Configure logging format
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # File handler
    file_handler = logging.FileHandler(log_file)
    file_handler.setFormatter(formatter)
    
    # Console handler
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    
    # Configure root logger
    logger = logging.getLogger(__name__)
    logger.setLevel(getattr(logging, log_level.upper()))
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    return logger

# --- Enhanced Barcode Scanner Class ---
class BarcodeScanner:
    def __init__(self, config: Config):
        self.config = config
        self.logger = setup_logging(config.log_level)
        self.excel_path = ""
        self.barcode_buffer = []
        self.last_input_time = None
        self.box_client = None
        self._lock = threading.Lock()
        self._running = False
        self.scan_count = 0  # Track number of scans for current file
        
        # Add upload queue and worker thread
        self.upload_queue = queue.Queue()
        self.upload_worker_running = False
        
        # Initialize components
        self._initialize_box_client()
        self._initialize_excel()
        self._setup_scheduler()
        self._start_upload_worker()
    
    def _initialize_box_client(self):
        """Initialize Box client with retry logic"""
        for attempt in range(self.config.max_retries):
            try:
                auth = JWTAuth.from_settings_file(self.config.box_config_path)
                self.box_client = Client(auth)
                # Test the connection
                user = self.box_client.user().get()
                self.logger.info(f"Box client initialized successfully for user: {user.name}")
                return
            except Exception as e:
                self.logger.warning(f"Box client initialization attempt {attempt + 1} failed: {e}")
                if attempt < self.config.max_retries - 1:
                    time.sleep(self.config.retry_delay)
        
        self.logger.error("Failed to initialize Box client after all retries")
        self.box_client = None
    
    def _initialize_excel(self):
        """Initialize Excel file with error handling"""
        today_str = datetime.now().strftime("%Y-%m-%d")
        self.excel_path = self.config.base_dir / f"scanned_barcodes_{today_str}.xlsx"
        
        try:
            if not self.excel_path.exists():
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Barcode Scans"
                ws.append(["Barcode", "Timestamp", "Status"])
                wb.save(self.excel_path)
                self.logger.info(f"Created new Excel file: {self.excel_path}")
                self.scan_count = 0  # Reset scan count for new file
            else:
                # Validate existing file and count existing scans
                wb = openpyxl.load_workbook(self.excel_path)
                ws = wb.active
                # Count rows minus header
                self.scan_count = max(0, ws.max_row - 1)
                self.logger.info(f"Using existing Excel file: {self.excel_path} with {self.scan_count} existing scans")
        except Exception as e:
            self.logger.error(f"Failed to initialize Excel file: {e}")
            raise
    
    def _setup_scheduler(self):
        """Set up daily rotation schedule"""
        schedule.every().day.at("00:01").do(self._rotate_excel)
        self.logger.info("Scheduled daily file rotation at 00:01")
    
    def _start_upload_worker(self):
        """Start persistent upload worker thread"""
        self.upload_worker_running = True
        worker_thread = threading.Thread(target=self._upload_worker, daemon=False)
        worker_thread.start()
        self.logger.info("Upload worker thread started")
    
    def _upload_worker(self):
        """Persistent worker thread that processes upload queue"""
        self.logger.info("Upload worker started and waiting for files...")
        
        while self.upload_worker_running:
            try:
                # Wait for file to upload (with timeout to check if we should stop)
                try:
                    file_path = self.upload_queue.get(timeout=5.0)
                except queue.Empty:
                    continue
                
                if file_path is None:  # Shutdown signal
                    break
                
                self.logger.info(f"Upload worker processing: {file_path}")
                
                # Attempt upload with retries
                success = self._upload_with_retries(file_path)
                
                if success:
                    self.logger.info(f"Successfully uploaded: {file_path}")
                else:
                    self.logger.error(f"Failed to upload after all retries: {file_path}")
                
                # Mark task as done
                self.upload_queue.task_done()
                
            except Exception as e:
                self.logger.error(f"Upload worker error: {e}")
                time.sleep(1)  # Brief pause before continuing
        
        self.logger.info("Upload worker stopped")
    
    def _upload_with_retries(self, file_path: Path, max_retries: int = None) -> bool:
        """Upload file with retry logic and proper logging"""
        if max_retries is None:
            max_retries = self.config.max_retries
        
        for attempt in range(max_retries):
            try:
                self.logger.info(f"Upload attempt {attempt + 1}/{max_retries} for {file_path.name}")
                
                # Check if file still exists and has data
                if not file_path.exists():
                    self.logger.warning(f"File no longer exists: {file_path}")
                    return False
                
                if not self._has_scan_data(file_path):
                    self.logger.info(f"File {file_path} has no scan data, deleting instead of uploading")
                    try:
                        file_path.unlink()
                        self.logger.info(f"Deleted empty file: {file_path}")
                    except Exception as e:
                        self.logger.warning(f"Could not delete empty file: {e}")
                    return True  # Consider this "successful" since we handled it
                
                # Attempt the upload
                success = self._upload_to_box(file_path)
                
                if success:
                    self.logger.info(f"Upload successful on attempt {attempt + 1}: {file_path.name}")
                    return True
                else:
                    self.logger.warning(f"Upload failed on attempt {attempt + 1}: {file_path.name}")
                    
            except Exception as e:
                self.logger.error(f"Upload attempt {attempt + 1} exception: {e}")
            
            # Wait before retry (except on last attempt)
            if attempt < max_retries - 1:
                wait_time = self.config.retry_delay * (attempt + 1)  # Exponential backoff
                self.logger.info(f"Waiting {wait_time} seconds before retry...")
                time.sleep(wait_time)
        
        self.logger.error(f"All upload attempts failed for: {file_path.name}")
        return False
    
    def _queue_file_for_upload(self, file_path: Path):
        """Add file to upload queue for processing by worker thread"""
        try:
            self.upload_queue.put(file_path, timeout=1.0)
            self.logger.info(f"Queued file for upload: {file_path.name}")
        except queue.Full:
            self.logger.error(f"Upload queue is full, cannot queue: {file_path.name}")
    
    def _wait_for_uploads_complete(self, timeout: int = 30):
        """Wait for all queued uploads to complete"""
        self.logger.info("Waiting for queued uploads to complete...")
        
        try:
            # Wait for queue to be empty
            start_time = time.time()
            while not self.upload_queue.empty():
                if time.time() - start_time > timeout:
                    self.logger.warning(f"Upload queue timeout after {timeout} seconds")
                    break
                time.sleep(1)
            
            # Wait for any in-progress uploads
            self.upload_queue.join()
            self.logger.info("All queued uploads completed")
            
        except Exception as e:
            self.logger.error(f"Error waiting for uploads: {e}")
      
    def _validate_barcode(self, barcode: str) -> bool:
        """Validate barcode format (customize as needed)"""
        if not barcode or len(barcode.strip()) == 0:
            return False
        
        # Add your barcode validation logic here
        # Example: check length, format, etc.
        if len(barcode.strip()) < 3:  # Minimum length check
            return False
        
        return True
    
    def _has_scan_data(self, file_path: Path) -> bool:
        """Check if Excel file has any scan data (more than just headers)"""
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            # Check if there are more rows than just the header
            return ws.max_row > 1
        except Exception as e:
            self.logger.error(f"Error checking scan data in {file_path}: {e}")
            return False
    
    def _log_barcode(self, barcode: str, status: str = "SUCCESS"):
        """Log barcode to Excel with thread safety"""
        with self._lock:
            try:
                wb = openpyxl.load_workbook(self.excel_path)
                ws = wb.active
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ws.append([barcode, timestamp, status])
                wb.save(self.excel_path)
                self.scan_count += 1  # Increment scan count
                self.logger.info(f"Logged barcode: {barcode} (Status: {status}). Total scans: {self.scan_count}")
            except InvalidFileException as e:
                self.logger.error(f"Excel file corrupted: {e}")
                self._handle_file_corruption(barcode)
            except Exception as e:
                self.logger.error(f"Failed to log barcode: {e}")
                self._handle_file_corruption(barcode)
    
    def _handle_file_corruption(self, barcode: str):
        """Handle corrupted Excel files"""
        timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        corrupted_file = str(self.excel_path).replace(".xlsx", f"_CORRUPTED_{timestamp_str}.xlsx")
        
        try:
            if self.excel_path.exists():
                os.rename(self.excel_path, corrupted_file)
                self.logger.warning(f"Renamed corrupt file to: {corrupted_file}")
        except Exception as e:
            self.logger.error(f"Could not rename corrupted file: {e}")
        
        # Create new file and log the current barcode
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Barcode Scans"
            ws.append(["Barcode", "Timestamp", "Status"])
            ws.append([barcode, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "RECOVERED"])
            wb.save(self.excel_path)
            self.scan_count = 1  # Reset scan count with the recovered barcode
            self.logger.info(f"Created new Excel file after corruption: {self.excel_path}")
        except Exception as e:
            self.logger.error(f"Failed to create new Excel file: {e}")
    
    def _upload_to_box(self, file_path: Path) -> bool:
        """Upload file to Box with improved error handling and logging"""
        if not file_path.exists():
            self.logger.warning(f"File not found for upload: {file_path}")
            return False
        
        if not self.box_client:
            self.logger.error("Box client not available for upload")
            return False
        
        try:
            folder = self.box_client.folder(self.config.target_folder_id)
            file_name = file_path.name
            
            self.logger.info(f"Starting Box upload: {file_name} ({file_path.stat().st_size} bytes)")
            
            # Check if file already exists
            try:
                items = folder.get_items()
                for item in items:
                    if item.name == file_name:
                        self.logger.info(f"File {file_name} already exists in Box (ID: {item.id})")
                        return True  # Consider this successful
            except Exception as e:
                self.logger.warning(f"Could not check existing files: {e}")
            
            # Upload the file
            with open(file_path, 'rb') as file_stream:
                uploaded_file = folder.upload_stream(file_stream, file_name)
            
            self.logger.info(f"Successfully uploaded {file_name} as Box file ID: {uploaded_file.id}")
            return True
            
        except BoxAPIException as e:
            self.logger.error(f"Box API error uploading {file_path.name}: {e}")
            if "invalid_grant" in str(e).lower():
                self.logger.error("Box JWT credentials may be expired or invalid")
            elif "not_found" in str(e).lower():
                self.logger.error("Box folder not found - check TARGET_FOLDER_ID")
            elif "insufficient_scope" in str(e).lower():
                self.logger.error("Box app lacks required permissions")
            return False
            
        except Exception as e:
            self.logger.error(f"Unexpected error uploading {file_path.name}: {e}")
            return False
    
    def _add_folder_collaborator(self):
        """Add collaborator to target folder"""
        if not self.box_client:
            self.logger.error("Box client not available for adding collaborator")
            return
        
        payload = {
            "item": {"type": "folder", "id": self.config.target_folder_id},
            "accessible_by": {"type": "user", "login": self.config.collaborator_email},
            "role": "co-owner"
        }
        
        try:
            response = self.box_client.session.post(
                "https://api.box.com/2.0/collaborations", 
                json=payload
            )
            if response.status_code in [200, 201, 202]:
                self.logger.info(f"Added collaborator {self.config.collaborator_email}")
            elif response.status_code == 400 and "user_already_collaborator" in response.text:
                self.logger.info(f"Collaborator {self.config.collaborator_email} already exists")
            else:
                self.logger.warning(f"Failed to add collaborator: {response.status_code} - {response.text}")
        except Exception as e:
            self.logger.error(f"Exception adding collaborator: {e}")
    
    def _rotate_excel(self):
        """Improved rotation with reliable upload queueing"""
        self.logger.info("Starting daily file rotation...")
        
        # Determine yesterday's file path
        yesterday_str = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        yesterday_file = self.config.base_dir / f"scanned_barcodes_{yesterday_str}.xlsx"
        
        # Handle existing current file
        current_file_handled = False
        if self.excel_path.exists():
            try:
                # Check if current file has scan data
                if self._has_scan_data(self.excel_path):
                    # Rename current file to yesterday's name if it doesn't exist
                    if not yesterday_file.exists():
                        self.excel_path.rename(yesterday_file)
                        self.logger.info(f"Renamed current file to: {yesterday_file}")
                        current_file_handled = True
                    else:
                        self.logger.warning(f"Yesterday file already exists: {yesterday_file}")
                else:
                    # Current file is empty, just delete it
                    self.excel_path.unlink()
                    self.logger.info(f"Deleted empty current file: {self.excel_path}")
                    current_file_handled = True
                    
            except Exception as e:
                self.logger.error(f"Error handling current file during rotation: {e}")
        
        # Queue yesterday's file for upload if it exists
        if yesterday_file.exists():
            self.logger.info(f"Queueing yesterday's file for upload: {yesterday_file}")
            self._queue_file_for_upload(yesterday_file)
        
        # Initialize new file for today
        try:
            self._initialize_excel()
            self.logger.info("File rotation completed successfully")
        except Exception as e:
            self.logger.error(f"Failed to initialize new Excel file: {e}")
    
    def _check_barcode_timeout(self):
        """Check for barcode input timeout"""
        if (self.barcode_buffer and self.last_input_time and 
            time.time() - self.last_input_time > self.config.barcode_timeout):
            
            incomplete_barcode = "".join(self.barcode_buffer)
            self.logger.warning(f"Barcode input timeout, clearing buffer: {incomplete_barcode}")
            self._log_barcode(incomplete_barcode, "TIMEOUT")
            self.barcode_buffer = []
            self.last_input_time = None
    
    def _on_barcode_input(self, event):
        """Handle barcode input with timeout and validation"""
        if not self._running or event.event_type != keyboard.KEY_DOWN:
            return
        
        current_time = time.time()
        self.last_input_time = current_time
        
        if event.name == "enter":
            if self.barcode_buffer:
                barcode = "".join(self.barcode_buffer).replace("space", " ").strip()
                
                if self._validate_barcode(barcode):
                    self._log_barcode(barcode, "SUCCESS")
                else:
                    self.logger.warning(f"Invalid barcode format: {barcode}")
                    self._log_barcode(barcode, "INVALID")
                
                self.barcode_buffer = []
                self.last_input_time = None
        else:
            # Collect barcode characters
            if len(event.name) == 1 or event.name == "space":
                self.barcode_buffer.append(event.name)
                
                # Prevent buffer overflow
                if len(self.barcode_buffer) > 100:  # Reasonable limit
                    self.logger.warning("Barcode buffer overflow, clearing")
                    self.barcode_buffer = []
                    self.last_input_time = None
    
    def stop(self):
        """Enhanced stop method that ensures uploads complete"""
        self.logger.info("Stopping barcode scanner...")
        self._running = False
        keyboard.unhook_all()
        
        # Wait for any pending uploads to complete
        self._wait_for_uploads_complete(timeout=60)
        
        # Stop upload worker
        self.upload_worker_running = False
        try:
            self.upload_queue.put(None, timeout=1.0)  # Signal worker to stop
        except:
            pass
        
        # Handle any remaining file uploads on shutdown
        if hasattr(self, 'excel_path') and self.excel_path.exists() and self._has_scan_data(self.excel_path):
            self.logger.info("Uploading current file before shutdown...")
            self._upload_with_retries(self.excel_path)
        
        self.logger.info("Barcode scanner stopped")
    
    def run(self):
        """Main execution loop"""
        try:
            self._running = True
            
            # Add collaborator on startup
            self._add_folder_collaborator()
            
            # Set up keyboard listener
            keyboard.on_press(self._on_barcode_input)
            self.logger.info("Barcode scanner service is running...")
            self.logger.info("Press Ctrl+C to stop")
            
            # Main loop
            while self._running:
                schedule.run_pending()
                self._check_barcode_timeout()
                time.sleep(0.1)  # Reduced sleep for better responsiveness
                
        except KeyboardInterrupt:
            self.logger.info("Received interrupt signal, shutting down...")
            self.stop()
        except Exception as e:
            self.logger.error(f"Unexpected error in main loop: {e}")
            self.stop()
            raise

# --- Main Execution ---
def main():
    try:
        config = Config()
        scanner = BarcodeScanner(config)
        scanner.run()
    except Exception as e:
        print(f"Failed to start barcode scanner: {e}")
        return 1
    return 0

if __name__ == "__main__":
    exit(main())
